# -*- coding: utf-8 -*-
# @Time        : 2022/4/26 9:52
# @Author      : martin_wang
# @FileName    : test2.py
# @IDE         : PyCharm
# @Description : 数据分析

import csv
import os
import shutil

from utils.utils import clear_folder, code_sum
from openpyxl import load_workbook


class Aggregation(object):
    def __init__(self, shot_image_save_folder, aggregation_file_save_folder, header, attribute_encode):
        self.stock_type = None
        self.aggregation_folder = None
        self.shot_image_save_folder = shot_image_save_folder
        self.aggregation_file_save_folder = aggregation_file_save_folder
        self.header = header
        self.attribute_encode = attribute_encode

        self.tuple_dict = {}
        self.count_dict = {}

    def _aggregation(self):
        self.tuple_dict.clear()
        self.count_dict.clear()
        for txt_name in os.listdir(self.aggregation_folder):
            attribute, _ = os.path.splitext(txt_name)
            txt_path = os.path.join(self.aggregation_folder, txt_name)
            with open(txt_path, 'r') as f:
                stock_codes = f.readlines()
                for stock_code in stock_codes:
                    stock_code = stock_code.strip()
                    if stock_code not in self.tuple_dict.keys():
                        _tuple = [" " for _ in range(len(self.header))]
                        _tuple[self.attribute_encode['stock_code']] = stock_code + '\t'
                        self.tuple_dict[stock_code] = _tuple
                        self.count_dict[stock_code] = 1
                    else:
                        self.count_dict[stock_code] = self.count_dict[stock_code] + 1
                    self.tuple_dict[stock_code][self.attribute_encode[attribute]] = "√"

    def to_csv(self, stock_type, csv_path):
        self.aggregation_folder = os.path.join(self.aggregation_file_save_folder, stock_type)
        self._aggregation()

        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(self.header)
            for stock_code, _tuple in self.tuple_dict.items():
                if self.count_dict[stock_code] > 1:
                # if self.count_dict[stock_code] > 1 and attribute[self.attribute_encode['15']] != " ":
                    _tuple[self.attribute_encode['code_sum']] = str(code_sum(stock_code)) + '\t'
                    writer.writerow(_tuple)

    def to_excel(self, stock_type, excel_path):
        self.aggregation_folder = os.path.join(self.aggregation_file_save_folder, stock_type)
        self._aggregation()

        work_book = load_workbook(excel_path)

        old_sheets = work_book.sheetnames
        for sheet in old_sheets:
            work_book.remove(work_book[sheet])

        sheet = work_book.create_sheet("Sheet1")

        sheet.append(self.header)

        for stock_code, _tuple in self.tuple_dict.items():
            if self.count_dict[stock_code] > 1:
                # if self.count_dict[stock_code] > 1 and attribute[self.attribute_encode['15']] != " ":
                _tuple[self.attribute_encode['code_sum']] = str(code_sum(stock_code)) + '\t'
                sheet.append(_tuple)

        work_book.save(excel_path)
        work_book.close()

    def to_txt(self, stock_type, curve_type):
        self.aggregation_folder = os.path.join(self.aggregation_file_save_folder,stock_type)
        image_folder = os.path.join(self.shot_image_save_folder, stock_type)
        txt_path = os.path.join(self.aggregation_folder, "{}.txt".format(curve_type))
        image_folder_path = os.path.join(image_folder, curve_type)
        with open(txt_path, 'w') as f:
            for image_name in os.listdir(image_folder_path):
                stock_code, _ = os.path.splitext(image_name)

                f.write('\t' + stock_code + "\n")

        shutil.rmtree(image_folder_path)


if __name__ == "__main__":
    pass
