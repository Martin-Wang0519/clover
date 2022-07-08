import os
import shutil
import sys
import traceback
from datetime import date, timedelta
from queue import Queue
from time import time, sleep

from PIL import ImageGrab
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QMessageBox
from openpyxl import load_workbook, Workbook

from utils.utils import get_window_handle, show_window, get_set_OCR_path, \
    logger, screen_clicked, screen_double_clicked, find_indent_click, keyboard_input, ocr, clear_folder
from predict import Predict
from shot_image_process import ShotImageProcess
from windows.window_ui import Ui_MainWindow
from _email import _Email
from import_images import ImportImages
from aggregation import Aggregation
from train import Trainer
from config import settings
from apscheduler.schedulers.blocking import BlockingScheduler
from apscheduler.triggers.cron import CronTrigger
from nets.model import ResNet34 as myModel
import torch

import logging

logging.basicConfig()
logging.getLogger('apscheduler').setLevel(logging.ERROR)


class Window(Ui_MainWindow, QMainWindow):
    def __init__(self, parent):
        super(Window, self).__init__(parent)

        self.setupUi(self)

        self.pos_dataset_path = 'dataset/stock_data/stock_images/positive'
        self.neg_dataset_path = 'dataset/stock_data/stock_images/negative'

        if not os.path.exists(self.pos_dataset_path):
            os.makedirs(self.pos_dataset_path)

        if not os.path.exists(self.neg_dataset_path):
            os.makedirs(self.neg_dataset_path)

        self.pos_samples_num = len(os.listdir(self.pos_dataset_path))
        self.neg_samples_num = len(os.listdir(self.neg_dataset_path))
        self.posLcdNumber.display(str(self.pos_samples_num))
        self.negLcdNumber.display(str(self.neg_samples_num))

        self.weight_path = "model_data/model_weight_cuda.pth"

        self.current_stock_type = None
        self.current_curve_type = None
        self.stock_code_ocr_crop_zone = None

        self.shot_num = None
        self.shot_image_save_path = settings.get('screen_shot_info').get('shot_image_save_path')
        self.stock_type_shot_image_save_path = None
        self.curve_type_shot_image_save_path = None

        self.RMZ_crop_zone = eval(settings.get('screen_shot_info').get('RMZ_crop_zone'))
        self.resolution = eval(settings.get('screen_shot_info').get('resolution'))
        self.grab_zone = (0, 0, self.resolution[0], self.resolution[1])

        #################################
        self.shot_frequency = 3
        self.shotFrequencySpinBox.setValue(self.shot_frequency)
        self.shot_image_queue = Queue()
        self.shot_image_process_thread_num = 12

        self.confidence = 0.99

        self.device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
        self.model = myModel(classes_num=2).to(self.device)

        ##################################
        self.task_queue = Queue(1000)
        self.scheduler = BlockingScheduler(timezone="Asia/Shanghai")
        self.arrow = "up_arrow"
        self.handle = None

        self.mail = _Email(settings.get('mail_info').get('from').get('address'),
                           settings.get('mail_info').get('from').get('auth_code'),
                           settings.get('mail_info').get('from').get('smtp_server'),
                           None)

        for address in settings.get('mail_info').get('to'):
            self.mailEdit.append(address)

        self.agg = Aggregation('daily_screenshot',
                               settings.get('statistics_info').get('folder_path'),
                               settings.get('statistics_info').get('prediction_excel_info').get('header'),
                               settings.get('statistics_info').get('prediction_excel_info').get(
                                   'time_curve_type_encode'))

        self.trainer = Trainer(self.model, 'dataset/stock_data', self.weight_path)
        get_set_OCR_path()

        self.logger = logger()

        self.shot_image_process_threads = []

        self.test_mode = False

    def quit_CGS_exe(self):
        point = eval(settings.get('click_point_info').get('CGS_close'))
        screen_clicked(point)
        sleep(2)

        point = eval(settings.get('click_point_info').get('CGS_quit_button'))
        screen_clicked(point)

    def execute_CGS_exe(self):
        point = eval(settings.get('click_point_info').get('CGS_task_bar'))

        screen_clicked(point)
        sleep(5)
        screen_clicked((1267, 624))
        sleep(5)

        self.handle = get_window_handle('海王星')

        show_window(self.handle, True)
        sleep(5)

    @pyqtSlot()
    def on_addPosPicButton_clicked(self):
        positive_samples_directory = QFileDialog.getExistingDirectory(self, "选择背离图片文件夹", "")
        if positive_samples_directory is None or positive_samples_directory == "":
            return
        try:
            self.textBrowser.append("正在从{}中导入背离图片，请稍后...".format(positive_samples_directory))
            self.textBrowser.repaint()

            mover = ImportImages(positive_samples_directory, self.pos_dataset_path, self.RMZ_crop_zone)
            mover.import_from_image_folder()

            pre_num = self.pos_samples_num
            self.pos_samples_num = len(os.listdir(self.pos_dataset_path))
            self.posLcdNumber.display(str(self.pos_samples_num))
            message = "背离图片导入成功！一共导入{}张背离图片".format(self.pos_samples_num - pre_num)
            self.textBrowser.append(message)
            self.logger.info(message)
        except RuntimeError:
            message = '图片导入失败\n' + traceback.format_exc()
            self.textBrowser.append(message)
            self.logger.error(message)
        finally:
            return

    @pyqtSlot()
    def on_addNegPicButton_clicked(self):
        negative_samples_directory = QFileDialog.getExistingDirectory(self, "选择背非离图片文件夹", "")
        if negative_samples_directory is None or negative_samples_directory == "":
            return
        try:
            self.textBrowser.append("正在从{}中导入非背离图片，请稍后...".format(negative_samples_directory))
            self.textBrowser.repaint()

            mover = ImportImages(negative_samples_directory, self.neg_dataset_path, self.RMZ_crop_zone)
            mover.import_from_image_folder()

            pre_num = self.neg_samples_num
            self.neg_samples_num = len(os.listdir(self.neg_dataset_path))
            self.negLcdNumber.display(str(self.neg_samples_num))
            message = "图片导入成功！一共导入{}张非背离图片".format(self.neg_samples_num - pre_num)
            self.textBrowser.append(message)
            self.logger.info(message)
        except RuntimeError:
            message = '图片导入失败\n' + traceback.format_exc()
            self.textBrowser.append(message)
            self.logger.error(message)
        finally:
            return

    @pyqtSlot()
    def on_trainButton_clicked(self):
        self.textBrowser.append("正在训练，请稍等...")
        self.textBrowser.repaint()

        self.trainer.start()
        self.textBrowser.append("训练完毕！")

    @pyqtSlot()
    def on_screenPredictButton_clicked(self):
        self.textBrowser.append('程序已启动')
        self.textBrowser.repaint()
        self.generate_schedule()
        self.mail.to_address = self.mailEdit.toPlainText().split('\n')

        if self.user_selection_check() is False:
            return

        self.scheduler.start()

    @pyqtSlot()
    def on_testButton_clicked(self):
        self.test_mode = True
        self.mail.to_address = settings.get('test_info').get('mail_to')
        if self.user_selection_check() is False:
            return

        for task in settings.get('test_info').get('tasks'):
            if task == 'gang' or task == 'a':
                self.switch_stock_type(task)
            elif task == 'email':
                self.send_mail()

            elif task == 'execute_CGS_exe':
                self.execute_CGS_exe()

            elif task == 'quit_CGS_exe':
                self.quit_CGS_exe()

            else:
                self.switch_curve_type(task)

    def update_parameter_switch_stock_type(self):
        if self.test_mode is True:
            self.shot_num = settings.get('test_info').get('shot_num').get(self.current_stock_type)
            self.stock_type_shot_image_save_path = os.path.join(self.shot_image_save_path, 'test',
                                                                self.current_stock_type)
        else:
            self.shot_num = settings.get('screen_shot_info').get('shot_num').get(self.current_stock_type)
            self.stock_type_shot_image_save_path = os.path.join(self.shot_image_save_path, self.current_stock_type)

        self.stock_code_ocr_crop_zone = eval(settings.get('screen_shot_info').get('ocr_zone')
                                             .get('code')
                                             .get(self.current_stock_type))
        self.arrow = "up_arrow"

    def switch_stock_type(self, stock_type):
        self.current_stock_type = stock_type
        self.update_parameter_switch_stock_type()

        screen_clicked(settings.get('click_point_info').get('zi_xuan_gu'))

        if self.current_stock_type == "gang":
            screen_clicked(settings.get('click_point_info').get('gang_gu'))
            screen_clicked(settings.get('click_point_info').get('gang_gu_tong'))
            sleep(10)
            screen_double_clicked(settings.get('click_point_info').get('gang_first'))
            sleep(5)
            screen_clicked(settings.get('click_point_info').get('left_menu'))

        elif self.current_stock_type == "a":
            screen_clicked(settings.get('click_point_info').get('a_gu'))
            sleep(10)
            screen_double_clicked(settings.get('click_point_info').get('a_first'))
            sleep(2)
            find_indent_click(eval(settings.get('click_point_info').get('indent')))
            sleep(5)
            screen_clicked(settings.get('click_point_info').get('left_menu'))

    def switch_curve_type(self, curve_type):
        self.current_curve_type = curve_type
        point = eval(settings.get('click_point_info').get(str(self.current_curve_type)))
        screen_clicked(point)
        sleep(5)

        self.arrow = "down_arrow" if self.arrow == "up_arrow" else "up_arrow"
        self.curve_type_shot_image_save_path = os.path.join(self.stock_type_shot_image_save_path,
                                                            self.current_curve_type)

        if not os.path.exists(self.curve_type_shot_image_save_path):
            os.makedirs(self.curve_type_shot_image_save_path)
        else:
            clear_folder(self.curve_type_shot_image_save_path)

        # for t in self.shot_image_process_threads:
        #     t.set_save_path(self.curve_type_shot_image_save_path)

        self.screenshot_predict()

    def screenshot_predict(self):
        sleep_time = 1.0 / self.shot_frequency
        global t
        for _ in range(self.shot_image_process_thread_num):
            t = ShotImageProcess(self.shot_image_queue, self.curve_type_shot_image_save_path,
                                 self.stock_code_ocr_crop_zone,
                                 self.RMZ_crop_zone)
            t.setDaemon(False)
            self.shot_image_process_threads.append(t)
            t.start()

        t0 = time()
        for _ in range(self.shot_num):
            grab_image = ImageGrab.grab(self.grab_zone)
            self.shot_image_queue.put(grab_image)
            keyboard_input(settings.get('VK_CODE_info').get(self.arrow), sleep_time)

        print('截图后队列大小'+str(self.shot_image_queue.qsize()))
        self.shot_image_queue.join()
        print('结束后队列大小'+str(self.shot_image_queue.qsize()))
        for t in self.shot_image_process_threads:
            t.stop()
        print('线程数'+str(t.get_enumerate()))

        t1 = time()
        t_shot = t1 - t0

        print("截图用时{}s,共{}张图片".format(t_shot, len(os.listdir(self.curve_type_shot_image_save_path))))

        self.textBrowser.append("截图用时{}s,共{}张图片".format(t_shot, len(os.listdir(self.curve_type_shot_image_save_path))))
        self.textBrowser.repaint()

        p = Predict(self.model, self.weight_path, self.curve_type_shot_image_save_path, self.confidence)
        p.folder_predict()
        t2 = time()
        t_predict = t2 - t1

        self.textBrowser.append("预测用时{}s".format(t_predict))
        self.textBrowser.repaint()
        self.agg.to_txt(self.current_stock_type, self.current_curve_type, self.curve_type_shot_image_save_path)

    def send_mail(self):
        excel_path = settings.get('statistics_info').get('prediction_excel_info').get('path')
        self.agg.to_excel(self.current_stock_type, excel_path)
        self.textBrowser.append("正在发送邮件...")

        if self.current_stock_type == 'a':
            self.mail.send_email(self.current_stock_type,
                                 [settings.get('statistics_info').get('prediction_excel_info').get('path'),
                                  'statistics/a/15.txt'])
        else:
            self.mail.send_email(self.current_stock_type,
                                 [settings.get('statistics_info').get('prediction_excel_info').get('path')])
        self.textBrowser.append("邮件发送完毕")
        shutil.copyfile(excel_path, settings.get('statistics_info').get('yesterday_prediction_path'))

    def on_aButton_clicked(self):
        self.current_stock_type = 'a'
        self.shot_num = 4709

    def on_gangButton_clicked(self):
        self.current_stock_type = 'gang'
        self.shot_num = 715

    def user_selection_check(self):

        if len(self.mailEdit.toPlainText()) == 0:
            QMessageBox.critical(self, "错误", "请输入邮箱！")
            return False

        return True

    def generate_schedule(self):
        for corn, task in settings.get('tasks_info').items():
            if task == 'execute_CGS_exe':
                self.scheduler.add_job(self.execute_CGS_exe,
                                       CronTrigger.from_crontab(corn, timezone='Asia/Shanghai'))

            elif task == 'quit_CGS_exe':
                self.scheduler.add_job(self.quit_CGS_exe,
                                       CronTrigger.from_crontab(corn, timezone='Asia/Shanghai'))

            elif task == 'email':
                self.scheduler.add_job(self.send_mail,
                                       CronTrigger.from_crontab(corn, timezone='Asia/Shanghai'))

            elif task == 'gang' or task == 'a':
                self.scheduler.add_job(self.switch_stock_type,
                                       CronTrigger.from_crontab(corn, timezone='Asia/Shanghai'),
                                       args=[task])


            else:
                self.scheduler.add_job(self.switch_curve_type,
                                       CronTrigger.from_crontab(corn, timezone='Asia/Shanghai'),
                                       args=[task])

    @pyqtSlot()
    def on_winningCalculationButton_clicked(self):
        self.handle = get_window_handle('海王星')

        if self.handle is None:
            QMessageBox.critical(self, "错误", "请打开海王星股票软件！")
            return False

        show_window(self.handle, True)

        self.current_stock_type = 'a'
        self.switch_stock_type(self.current_stock_type)

        yesterday_prediction_excel_path = settings.get('statistics_info').get('yesterday_prediction_path')

        work_book = load_workbook(yesterday_prediction_excel_path)
        sheet = work_book.active

        stock_codes = [cell.value for cell in sheet['A']]

        stock_codes.remove('股票代码')

        if len(stock_codes) == 0:
            return

        sum_rise = 0
        winning_sum = 0
        count = 0
        for stock_code in stock_codes:
            for ch in stock_code.strip():
                keyboard_input(settings.get('VK_CODE_info').get(ch), 0.5)
            keyboard_input(settings.get('VK_CODE_info').get('enter'), 0.5)

            grab_image = ImageGrab.grab(self.grab_zone)
            rise = ocr(grab_image,
                       eval(settings.get('screen_shot_info').get('ocr_zone').get('rise').get('gang'))).rstrip('%')
            try:
                rise = float(rise)
                rise = rise / 100 if rise > 100 else rise
                print(stock_code, rise)
                winning_sum = winning_sum + 1 if rise > 0 else winning_sum

                sum_rise = sum_rise + rise

                count = count + 1
            except Exception:
                self.logger.error(traceback.format_exc())

        winning_rate = round(winning_sum / count * 100, 2)
        average_rise = round(sum_rise / count, 2)
        winning_excel_path = settings.get('statistics_info').get('winning_excel_info').get('path')
        if not os.path.exists(winning_excel_path):
            work_book = Workbook()
            sheet = work_book.active
            header = settings.get('statistics_info').get('winning_excel_info').get('header')
            sheet.append(header)
            work_book.save(winning_excel_path)
            work_book.close()

        today = date.today()
        back_days = -1 if today.isoweekday() != 1 else -3
        prediction_date = (date.today() + timedelta(days=back_days)).strftime("%Y-%m-%d")
        data = [prediction_date, winning_rate, average_rise]

        work_book = load_workbook(winning_excel_path)
        sheet = work_book.active

        sheet.append(data)

        work_book.save(winning_excel_path)
        work_book.close()

        self.handle = get_window_handle('股票预测')

        show_window(self.handle, False)

        self.textBrowser.append('胜率统计完成，上次预测准确率为{}%，平均涨点为{}%'.format(winning_rate, average_rise))

    def closeEvent(self, event):
        event.accept()
        sys.exit(0)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    mw = Window(None)
    mw.show()
    sys.exit(app.exec_())
